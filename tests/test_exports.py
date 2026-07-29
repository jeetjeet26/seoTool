import csv
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from modules.models import Finding, Severity
from worker.exports import generate_report_exports


class ExportTests(unittest.TestCase):
    def test_generates_csv_and_hyperlinked_excel_findings(self):
        finding = Finding(
            id="a" * 64,
            category="links",
            severity=Severity.HIGH,
            issue_type="client_error_4xx",
            page_url="https://example.com/",
            resource_url="https://example.com/missing",
            evidence='{"status_code":"404"}',
            recommendation="Update the link.",
            source_file="response_codes_client_error_(4xx).csv",
        )
        summary = {
            "finding_count": 1,
            "target_url": "https://example.com/",
            "target_location": "Dallas, Texas",
            "content_recommendations": [
                {
                    "url": "https://example.com/",
                    "current_title": "Old",
                    "proposed_title": "New",
                    "requires_human_review": True,
                }
            ],
            "page_experience": [
                {
                    "url": "https://example.com/",
                    "performance_score": 90,
                    "accessibility_score": 95,
                }
            ],
        }

        with tempfile.TemporaryDirectory() as directory:
            paths = generate_report_exports(
                "audit-1", [finding], summary, Path(directory)
            )
            csv_path = next(path for path in paths if path.suffix == ".csv")
            xlsx_path = next(path for path in paths if path.suffix == ".xlsx")

            with csv_path.open(encoding="utf-8-sig") as stream:
                rows = list(csv.reader(stream))
            workbook = load_workbook(xlsx_path)

        self.assertEqual(rows[1][3], "https://example.com/")
        self.assertIn("Introduction", workbook.sheetnames)
        self.assertIn("Title Tags", workbook.sheetnames)
        self.assertIn("Technical SEO", workbook.sheetnames)
        self.assertIn("Page Speed", workbook.sheetnames)
        self.assertIn("Glossary", workbook.sheetnames)
        # Proposed title lands in the Title Tags sheet.
        titles = workbook["Title Tags"]
        self.assertEqual(titles.cell(row=6, column=5).value, "New")
        # Technical rows keep occurrence counts and hyperlink the example URL.
        technical = workbook["Technical SEO"]
        self.assertEqual(technical.cell(row=6, column=4).value, 1)
        self.assertEqual(
            technical.cell(row=6, column=5).hyperlink.target,
            "https://example.com/",
        )


if __name__ == "__main__":
    unittest.main()
