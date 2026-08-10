import csv
import io
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from worker.export_profiles import (
    SeopressTemplateError,
    build_client_workbook,
    developer_compilation_csv,
    merge_seopress_template,
    metadata_review_csv,
    sanitize_cell,
)


class SeopressTemplateTests(unittest.TestCase):
    TEMPLATE = (
        "ID,url,seopress_titles_title,seopress_titles_desc,custom_field\n"
        "1,https://example.com/,Old title,Old description,keep-me\n"
        "2,https://example.com/other/,Other title,Other description,also-keep\n"
    ).encode("utf-8")

    def test_preserves_order_and_unknown_columns(self):
        merged = merge_seopress_template(
            self.TEMPLATE,
            {
                "https://example.com/": {
                    "title": "New title",
                    "meta_description": "New description",
                }
            },
        )
        rows = list(csv.reader(io.StringIO(merged)))
        self.assertEqual(
            rows[0],
            ["ID", "url", "seopress_titles_title", "seopress_titles_desc", "custom_field"],
        )
        self.assertEqual(rows[1][2], "New title")
        self.assertEqual(rows[1][3], "New description")
        self.assertEqual(rows[1][4], "keep-me")
        # Untouched row passes through byte-for-byte values.
        self.assertEqual(rows[2][2], "Other title")
        self.assertEqual(rows[2][4], "also-keep")
        self.assertEqual(len(rows), 3)

    def test_requires_url_column(self):
        template = b"ID,name\n1,Homepage\n"
        with self.assertRaises(SeopressTemplateError):
            merge_seopress_template(template, {})

    def test_requires_writable_column(self):
        template = b"url,extra\nhttps://example.com/,x\n"
        with self.assertRaises(SeopressTemplateError):
            merge_seopress_template(template, {})

    def test_rejects_non_utf8(self):
        with self.assertRaises(SeopressTemplateError):
            merge_seopress_template(b"\xff\xfe\x00bad", {})

    def test_formula_injection_is_neutralized(self):
        merged = merge_seopress_template(
            self.TEMPLATE,
            {
                "https://example.com/": {
                    "title": "=HYPERLINK(evil)",
                    "meta_description": "+SUM(1)",
                }
            },
        )
        rows = list(csv.reader(io.StringIO(merged)))
        self.assertEqual(rows[1][2], "'=HYPERLINK(evil)")
        self.assertEqual(rows[1][3], "'+SUM(1)")


class CsvProfileTests(unittest.TestCase):
    ITEM = {
        "url": "https://example.com/floor-plans/",
        "keywords": ["2 bedroom apartments"],
        "current_title": "Old",
        "proposed_title": "2 Bedroom Apartments in Dallas | Example",
        "current_meta_description": "Old description",
        "proposed_meta_description": "D" * 140,
        "current_h1": "Old H1",
        "proposed_h1": "New H1",
        "proposed_content": "New intro copy.",
    }

    def test_developer_compilation(self):
        output = developer_compilation_csv([self.ITEM])
        rows = list(csv.reader(io.StringIO(output)))
        self.assertEqual(rows[0][0], "PAGE NAME")
        self.assertEqual(rows[1][0], "Floor Plans")
        self.assertEqual(rows[1][2], self.ITEM["proposed_title"])
        self.assertEqual(rows[1][5], "New intro copy.")

    def test_metadata_review_lengths(self):
        output = metadata_review_csv([self.ITEM])
        rows = list(csv.reader(io.StringIO(output)))
        header = rows[0]
        row = dict(zip(header, rows[1]))
        self.assertEqual(row["Proposed Title"], self.ITEM["proposed_title"])
        self.assertEqual(
            int(row["Proposed Title Length"]), len(self.ITEM["proposed_title"])
        )
        self.assertEqual(int(row["Proposed Description Length"]), 140)

    def test_sanitize_cell(self):
        self.assertEqual(sanitize_cell("=cmd()"), "'=cmd()")
        self.assertEqual(sanitize_cell("safe"), "safe")
        self.assertEqual(sanitize_cell(12), 12)


class ClientWorkbookTests(unittest.TestCase):
    def test_only_populated_sheets_are_generated(self):
        workbook = build_client_workbook("Test Property")
        self.assertEqual(workbook.sheetnames, ["Introduction"])
        # No hidden sheets ever.
        self.assertTrue(
            all(workbook[name].sheet_state == "visible" for name in workbook.sheetnames)
        )

    def test_full_workbook_structure(self):
        workbook = build_client_workbook(
            "Test Property",
            keywords=[
                {
                    "keyword": "apartments dallas",
                    "position": 5,
                    "intent": "commercial",
                    "cpc": 1.2,
                    "volume": 800,
                    "difficulty": 40,
                    "assigned_page": "https://example.com/",
                }
            ],
            metadata_items=[CsvProfileTests.ITEM],
            onpage_items=[CsvProfileTests.ITEM],
            alt_text_items=[
                {
                    "page_url": "https://example.com/",
                    "image_url": "https://example.com/pool.jpg",
                    "proposed_alt_text": "Resort-style pool",
                }
            ],
            technical_rows=[
                {
                    "category": "links",
                    "issue": "Client Error 4xx",
                    "description": "Broken internal links",
                    "occurrences": 3,
                    "example_url": "https://example.com/missing",
                    "recommendation": "Update the links.",
                }
            ],
            page_experience=[
                {"url": "https://example.com/", "performance_score": 88, "accessibility_score": 95}
            ],
            recap_lines=["Pages crawled: 40"],
        )
        expected = [
            "Introduction",
            "Keyword Research",
            "Title Tags",
            "Description Tags",
            "H1 Tags",
            "On-Page SEO",
            "Technical SEO",
        ]
        self.assertEqual(workbook.sheetnames, expected)

        intro = workbook["Introduction"]
        self.assertEqual(intro["A1"].value, "Test Property")
        toc_values = [intro.cell(row=row, column=2).value for row in range(4, 15)]
        self.assertIn("Keyword Research", toc_values)
        self.assertIn("Technical SEO", toc_values)

        titles = workbook["Title Tags"]
        self.assertEqual(titles.cell(row=5, column=1).value, "URL")
        self.assertEqual(
            titles.cell(row=6, column=5).value, CsvProfileTests.ITEM["proposed_title"]
        )

        technical = workbook["Technical SEO"]
        self.assertEqual(technical.cell(row=6, column=4).value, 3)
        # URLs are hyperlinked.
        self.assertEqual(
            technical.cell(row=6, column=5).hyperlink.target,
            "https://example.com/missing",
        )

    def test_in_house_workbook_uses_condensed_treatment(self):
        workbook = build_client_workbook(
            "Test Property",
            keywords=[{"keyword": "apartments dallas"}],
            metadata_items=[CsvProfileTests.ITEM],
            technical_rows=[{"category": "links", "issue": "Broken link"}],
            report_variant="in_house",
        )
        self.assertEqual(
            workbook.sheetnames,
            ["Keyword Research", "SEO Treatment", "Technical SEO"],
        )
        self.assertEqual(
            workbook["SEO Treatment"].cell(row=6, column=1).value,
            CsvProfileTests.ITEM["url"],
        )

    def test_workbook_saves_and_reloads(self):
        workbook = build_client_workbook(
            "Test Property",
            metadata_items=[CsvProfileTests.ITEM],
        )
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "report.xlsx"
            workbook.save(path)
            reloaded = load_workbook(path)
        self.assertIn("Title Tags", reloaded.sheetnames)

    def test_formula_values_are_sanitized_in_tables(self):
        workbook = build_client_workbook(
            "Test Property",
            metadata_items=[{**CsvProfileTests.ITEM, "proposed_title": "=EVIL()"}],
        )
        titles = workbook["Title Tags"]
        self.assertEqual(titles.cell(row=6, column=5).value, "'=EVIL()")


if __name__ == "__main__":
    unittest.main()
