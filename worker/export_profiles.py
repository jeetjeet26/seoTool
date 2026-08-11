"""Named, testable export profiles for audit and tool-run deliverables.

Profiles:
- ``seopress_csv``: merge approved metadata into an uploaded SEOPress template
  CSV, preserving row order and every unrecognized column byte-for-byte.
- ``internal_findings_csv``: the flat internal findings export.
- ``developer_compilation_csv``: one-sheet handoff a web developer can work
  from (page name, URL, title, description, H1, on-page copy).
- ``client_workbook``: multi-sheet Excel report modeled on the agency's
  reference workbook (Introduction/TOC, Keyword Research, Title Tags,
  Description Tags, H1 Tags, On-Page SEO, Alt Text, Technical SEO, Page Speed,
  Program Recap, Glossary). Only populated sheets and rows are generated.

All CSV/XLSX cell values are hardened against spreadsheet formula injection.
"""

from __future__ import annotations

import csv
import io
from datetime import date
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

HEADER_FILL = PatternFill("solid", fgColor="176B52")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(size=16, bold=True, color="176B52")

# Columns we are allowed to write into an uploaded SEOPress template. Every
# other column is preserved untouched.
SEOPRESS_URL_COLUMNS = {"url", "permalink", "address", "page url"}
SEOPRESS_TITLE_COLUMNS = {"seopress_titles_title", "meta title", "seo title", "title"}
SEOPRESS_DESCRIPTION_COLUMNS = {
    "seopress_titles_desc",
    "meta description",
    "seo description",
    "description",
}

GLOSSARY: tuple[tuple[str, str], ...] = (
    ("Title Tag", "The clickable headline of a search result. It should accurately and concisely describe the page in roughly 60 characters."),
    ("Meta Description", "The short paragraph under a search result headline. It does not directly affect rankings but strongly affects click-through rate."),
    ("H1 / Headings", "The visible headline structure of a page (H1, H2, H3). The H1 should describe the page and appear exactly once."),
    ("Canonical Tag", "A tag that tells search engines which URL is the primary version of a page so duplicates do not compete with each other."),
    ("Crawl", "The process where a search engine robot reads a site to discover and evaluate its pages."),
    ("Indexed Pages", "Pages a search engine has stored and can show in results."),
    ("XML Sitemap", "A machine-readable file listing the URLs a site wants search engines to crawl and index."),
    ("Backlink", "A link from another website to yours. Quality backlinks are a meaningful ranking signal."),
    ("Image Alt Text", "A text description of an image used by screen readers and search engines."),
    ("Redirect", "A rule that forwards one URL to another. Internal links should point at final destinations, not redirects."),
    ("Schema / Structured Data", "Code added to a page that describes its content so search engines can show richer results."),
    ("Core Web Vitals", "Google's page-experience metrics covering loading speed, interactivity, and visual stability."),
    ("SERP", "Search Engine Results Page - the page of results a search engine returns for a query."),
    ("Keyword Difficulty", "An estimate (0-100) of how hard it is to rank on page one for a keyword."),
    ("Search Volume", "The average number of monthly searches for a keyword."),
)


def sanitize_cell(value: Any) -> Any:
    """Neutralize spreadsheet formula injection for untrusted text."""
    if isinstance(value, str) and value and value[0] in ("=", "+", "-", "@", "\t"):
        return "'" + value
    return value


def _sanitize_row(row: Iterable[Any]) -> list[Any]:
    return [sanitize_cell(value) for value in row]


# ---------------------------------------------------------------------------
# SEOPress template merge
# ---------------------------------------------------------------------------


class SeopressTemplateError(ValueError):
    """Raised when the uploaded template cannot be safely merged."""


def merge_seopress_template(
    template_bytes: bytes,
    metadata_by_url: dict[str, dict[str, str]],
    max_rows: int = 20000,
) -> str:
    """Fill approved titles/descriptions into an uploaded SEOPress CSV.

    ``metadata_by_url`` maps a page URL to ``{"title": ..., "meta_description": ...}``.
    Rows whose URL has no approved metadata are passed through untouched, as is
    every column we do not recognize. Row order is preserved.
    """

    try:
        text = template_bytes.decode("utf-8-sig")
    except UnicodeDecodeError as exc:
        raise SeopressTemplateError("The template must be UTF-8 encoded") from exc

    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    if not rows:
        raise SeopressTemplateError("The template CSV is empty")
    if len(rows) - 1 > max_rows:
        raise SeopressTemplateError(f"The template exceeds {max_rows} rows")

    header = rows[0]
    normalized = [column.strip().lower() for column in header]
    url_index = _find_column(normalized, SEOPRESS_URL_COLUMNS)
    title_index = _find_column(normalized, SEOPRESS_TITLE_COLUMNS)
    description_index = _find_column(normalized, SEOPRESS_DESCRIPTION_COLUMNS)

    if url_index is None:
        raise SeopressTemplateError(
            "The template needs a URL column (url, permalink, or address)"
        )
    if title_index is None and description_index is None:
        raise SeopressTemplateError(
            "The template needs a title or description column to fill"
        )

    lookup = {_url_key(url): values for url, values in metadata_by_url.items()}

    output = io.StringIO()
    writer = csv.writer(output, lineterminator="\r\n")
    writer.writerow(header)
    for row in rows[1:]:
        padded = row + [""] * (len(header) - len(row))
        key = _url_key(padded[url_index]) if url_index < len(padded) else ""
        approved = lookup.get(key)
        if approved:
            if title_index is not None and approved.get("title"):
                padded[title_index] = str(sanitize_cell(approved["title"]))
            if description_index is not None and approved.get("meta_description"):
                padded[description_index] = str(
                    sanitize_cell(approved["meta_description"])
                )
        writer.writerow(padded[: len(header)])
    return output.getvalue()


def _find_column(normalized_header: list[str], aliases: set[str]) -> int | None:
    for index, name in enumerate(normalized_header):
        if name in aliases:
            return index
    return None


def _url_key(url: str) -> str:
    return (url or "").strip().rstrip("/").lower()


# ---------------------------------------------------------------------------
# Internal CSV profiles
# ---------------------------------------------------------------------------


def internal_findings_csv(findings: Iterable[Any]) -> str:
    """The flat internal findings export (severity-ordered as provided)."""
    output = io.StringIO()
    writer = csv.writer(output, lineterminator="\r\n")
    writer.writerow(
        [
            "Severity",
            "Category",
            "Issue Type",
            "Page URL",
            "Resource URL",
            "Evidence",
            "Recommendation",
            "Source File",
        ]
    )
    for finding in findings:
        writer.writerow(
            _sanitize_row(
                [
                    finding.severity.value,
                    finding.category,
                    finding.issue_type,
                    finding.page_url,
                    finding.resource_url,
                    finding.evidence,
                    finding.recommendation,
                    finding.source_file,
                ]
            )
        )
    return output.getvalue()


def developer_compilation_csv(items: Iterable[dict]) -> str:
    """Single handoff sheet for the web developer, one row per page."""
    output = io.StringIO()
    writer = csv.writer(output, lineterminator="\r\n")
    writer.writerow(
        ["PAGE NAME", "URL", "TITLE TAG", "META DESCRIPTION", "H1", "ON PAGE OPTIMIZATION"]
    )
    for item in items:
        writer.writerow(
            _sanitize_row(
                [
                    item.get("page_name") or _page_name(item.get("url", "")),
                    item.get("url", ""),
                    item.get("proposed_title") or item.get("title", ""),
                    item.get("proposed_meta_description")
                    or item.get("meta_description", ""),
                    item.get("proposed_h1") or item.get("h1", ""),
                    item.get("proposed_content") or item.get("content", ""),
                ]
            )
        )
    return output.getvalue()


def metadata_review_csv(items: Iterable[dict]) -> str:
    """Normal CSV: current vs approved values with lengths, one row per page."""
    output = io.StringIO()
    writer = csv.writer(output, lineterminator="\r\n")
    writer.writerow(
        [
            "URL",
            "Keywords",
            "Current Title",
            "Proposed Title",
            "Proposed Title Length",
            "Current Meta Description",
            "Proposed Meta Description",
            "Proposed Description Length",
            "Current H1",
            "Proposed H1",
        ]
    )
    for item in items:
        proposed_title = item.get("proposed_title") or item.get("title") or ""
        proposed_description = (
            item.get("proposed_meta_description") or item.get("meta_description") or ""
        )
        writer.writerow(
            _sanitize_row(
                [
                    item.get("url", ""),
                    "; ".join(item.get("keywords") or []),
                    item.get("current_title", ""),
                    proposed_title,
                    len(proposed_title),
                    item.get("current_meta_description", ""),
                    proposed_description,
                    len(proposed_description),
                    item.get("current_h1", ""),
                    item.get("proposed_h1") or item.get("h1") or "",
                ]
            )
        )
    return output.getvalue()


def _page_name(url: str) -> str:
    path = url.rstrip("/").rsplit("/", 1)[-1] if "//" in url else url
    if not path or "." in path or path.startswith("http"):
        return "Home Page"
    return path.replace("-", " ").replace("_", " ").title()


# ---------------------------------------------------------------------------
# Client workbook (Alexan-style, clean)
# ---------------------------------------------------------------------------


def build_client_workbook(
    property_name: str,
    keywords: list[dict] | None = None,
    metadata_items: list[dict] | None = None,
    onpage_items: list[dict] | None = None,
    alt_text_items: list[dict] | None = None,
    technical_rows: list[dict] | None = None,
    page_experience: list[dict] | None = None,
    recap_lines: list[str] | None = None,
    report_variant: str = "full_client",
) -> Workbook:
    """Build the multi-sheet client report workbook.

    Only sections with data get a sheet; every sheet contains only populated
    rows, valid values, and no hidden or duplicated tabs.
    """

    workbook = Workbook()
    intro = workbook.active
    intro.title = "Introduction"

    sections: list[str] = []

    if keywords:
        sheet = workbook.create_sheet("Keyword Research")
        _sheet_title(sheet, property_name, "Keyword Research")
        _table(
            sheet,
            ["Keywords", "Ranking", "Intent", "CPC", "Volume", "Difficulty", "Target Page"],
            [
                [
                    row.get("keyword", ""),
                    row.get("position") or "-",
                    (row.get("intent") or "commercial")[:1],
                    row.get("cpc") or "n/a",
                    row.get("volume") or "n/a",
                    row.get("difficulty") or "n/a",
                    row.get("assigned_page", ""),
                ]
                for row in keywords
            ],
        )
        sections.append("Keyword Research")

    if metadata_items and report_variant == "full_client":
        revised = date.today().strftime("%-m/%-d/%y")
        titles = workbook.create_sheet("Title Tags")
        _sheet_title(titles, property_name, "Title Tags")
        _table(
            titles,
            ["URL", "Keywords", "Current Title", "Length", "Proposed Title", "Length", "Date Revised"],
            [
                [
                    item.get("url", ""),
                    "; ".join(item.get("keywords") or []),
                    item.get("current_title", ""),
                    len(item.get("current_title") or ""),
                    item.get("proposed_title") or item.get("title") or "",
                    len(item.get("proposed_title") or item.get("title") or ""),
                    revised,
                ]
                for item in metadata_items
            ],
        )
        sections.append("Page Titles")

        descriptions = workbook.create_sheet("Description Tags")
        _sheet_title(descriptions, property_name, "Description Tags")
        _table(
            descriptions,
            [
                "URL",
                "Keywords",
                "Current Description",
                "Length",
                "Proposed Description",
                "Length",
                "Date Revised",
            ],
            [
                [
                    item.get("url", ""),
                    "; ".join(item.get("keywords") or []),
                    item.get("current_meta_description", ""),
                    len(item.get("current_meta_description") or ""),
                    item.get("meta_description")
                    or item.get("proposed_meta_description")
                    or "",
                    len(
                        item.get("meta_description")
                        or item.get("proposed_meta_description")
                        or ""
                    ),
                    revised,
                ]
                for item in metadata_items
            ],
        )
        sections.append("Page Descriptions")

        h1_rows = [
            [
                item.get("url", ""),
                "; ".join(item.get("keywords") or []),
                item.get("current_h1", ""),
                item.get("proposed_h1") or item.get("h1") or "",
            ]
            for item in metadata_items
            if item.get("proposed_h1") or item.get("h1") or item.get("current_h1")
        ]
        if h1_rows:
            h1_sheet = workbook.create_sheet("H1 Tags")
            _sheet_title(h1_sheet, property_name, "H1 Tags")
            h1_rows = [
                [
                    row[0],
                    row[1],
                    row[2],
                    len(row[2] or ""),
                    row[3],
                    len(row[3] or ""),
                    revised,
                ]
                for row in h1_rows
            ]
            _table(h1_sheet, ["URL", "Keywords", "Current H1", "Length", "Proposed H1", "Length", "Date Revised"], h1_rows)
            sections.append("H1 Tags")

    if metadata_items and report_variant == "in_house":
        treatment = workbook.create_sheet("SEO Treatment")
        _sheet_title(treatment, property_name, "SEO Treatment")
        _table(
            treatment,
            [
                "URL",
                "Keywords",
                "Proposed Title",
                "Length Title",
                "Proposed Description",
                "Length Description",
                "Proposed H1",
                "Length H1",
            ],
            [
                [
                    item.get("url", ""),
                    "; ".join(item.get("keywords") or []),
                    item.get("proposed_title") or item.get("current_title") or "",
                    len(item.get("proposed_title") or item.get("current_title") or ""),
                    item.get("proposed_meta_description")
                    or item.get("current_meta_description")
                    or "",
                    len(
                        item.get("proposed_meta_description")
                        or item.get("current_meta_description")
                        or ""
                    ),
                    item.get("proposed_h1") or item.get("current_h1") or "",
                    len(item.get("proposed_h1") or item.get("current_h1") or ""),
                ]
                for item in metadata_items
            ],
        )
        sections.append("SEO Treatment")

    if onpage_items and report_variant == "full_client":
        onpage = workbook.create_sheet("On-Page SEO")
        _sheet_title(onpage, property_name, "On-Page SEO Recommendations")
        row_index = 5
        for item in onpage_items:
            entries = [
                ("Page:", item.get("url", "")),
                ("Targeted Keyword(s):", "; ".join(item.get("keywords") or [])),
                (
                    "Recommendation Type:",
                    "New paragraph block"
                    if item.get("content_action") == "new_block"
                    else "Light paragraph rewrite",
                ),
                ("Original Copy:", item.get("current_body_text") or ""),
                ("Proposed Copy:", item.get("proposed_content") or item.get("content", "")),
            ]
            for label, value in entries:
                onpage.cell(row=row_index, column=1, value=label).font = Font(bold=True)
                onpage.cell(row=row_index, column=2, value=sanitize_cell(value))
                row_index += 1
            row_index += 1
        onpage.column_dimensions["A"].width = 24
        onpage.column_dimensions["B"].width = 110
        for row in onpage.iter_rows(min_row=5, max_col=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        sections.append("On Page SEO")

    if alt_text_items and report_variant not in {"full_client", "in_house"}:
        alt_sheet = workbook.create_sheet("Alt Text")
        _sheet_title(alt_sheet, property_name, "Image Alt Text Recommendations")
        _table(
            alt_sheet,
            ["Page", "Image", "Current Alt Text", "Proposed Alt Text", "Length"],
            [
                [
                    item.get("page_url", ""),
                    item.get("image_url", ""),
                    item.get("current_alt_text") or "-",
                    item.get("proposed_alt_text") or item.get("alt_text") or "",
                    len(item.get("proposed_alt_text") or item.get("alt_text") or ""),
                ]
                for item in alt_text_items
            ],
        )
        sections.append("Image Alt Text")

    if technical_rows:
        technical = workbook.create_sheet("Technical SEO")
        _sheet_title(technical, property_name, "Technical SEO Recommendations")
        _table(
            technical,
            [
                "Type",
                "Issue",
                "Description",
                "Occurrences",
                "Example Location",
                "Recommendation",
            ],
            [
                [
                    row.get("category", ""),
                    row.get("issue", ""),
                    row.get("description", ""),
                    row.get("occurrences", 1),
                    row.get("example_url", ""),
                    row.get("recommendation", ""),
                ]
                for row in technical_rows
            ],
        )
        sections.append("Technical SEO")

    if page_experience and report_variant not in {"full_client", "in_house"}:
        speed = workbook.create_sheet("Page Speed")
        _sheet_title(speed, property_name, "Page Speed & Accessibility")
        _table(
            speed,
            ["URL", "Performance Score", "Accessibility Score"],
            [
                [
                    row.get("url", ""),
                    row.get("performance_score", ""),
                    row.get("accessibility_score", ""),
                ]
                for row in page_experience
            ],
        )
        sections.append("Page Speed")

    if recap_lines and report_variant not in {"full_client", "in_house"}:
        recap = workbook.create_sheet("Program Recap")
        _sheet_title(recap, property_name, "Program Recap")
        row_index = 5
        for line in recap_lines:
            recap.cell(row=row_index, column=1, value=sanitize_cell(line))
            row_index += 1
        recap.column_dimensions["A"].width = 120
        sections.append("Program Recap")

    if report_variant not in {"full_client", "in_house"}:
        glossary = workbook.create_sheet("Glossary")
        _sheet_title(glossary, property_name, "Glossary of Terms")
        row_index = 5
        for term, definition in GLOSSARY:
            cell = glossary.cell(row=row_index, column=1, value=f"{term} - {definition}")
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            row_index += 2
        glossary.column_dimensions["A"].width = 120
        sections.append("Glossary")

    # Introduction / table of contents last, once sections are known.
    intro.cell(row=1, column=1, value=property_name).font = TITLE_FONT
    intro.cell(row=4, column=2, value="Table Of Contents").font = Font(bold=True, size=12)
    for offset, section in enumerate(sections):
        intro.cell(row=5 + offset, column=2, value=section)
    intro.column_dimensions["A"].width = 8
    intro.column_dimensions["B"].width = 48
    if report_variant == "in_house":
        workbook.remove(intro)

    return workbook


def _sheet_title(sheet, property_name: str, section: str) -> None:
    sheet.cell(row=1, column=1, value=f"{property_name} / {section}").font = TITLE_FONT
    sheet.cell(
        row=3,
        column=1,
        value="These are edits that should be made by your web developer.",
    ).font = Font(italic=True, color="666666")


def _table(sheet, headers: list[str], rows: list[list[Any]], start_row: int = 5) -> None:
    for column_index, header in enumerate(headers, start=1):
        cell = sheet.cell(row=start_row, column=column_index, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for row_offset, row in enumerate(rows, start=1):
        for column_index, value in enumerate(_sanitize_row(row), start=1):
            cell = sheet.cell(
                row=start_row + row_offset, column=column_index, value=value
            )
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if isinstance(value, str) and value.startswith("http"):
                cell.hyperlink = value
                cell.style = "Hyperlink"
    for column_cells in sheet.iter_cols(
        min_row=start_row, max_row=start_row + len(rows), max_col=len(headers)
    ):
        letter = column_cells[0].column_letter
        longest = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[letter].width = min(70, max(14, longest + 2))
    sheet.freeze_panes = sheet.cell(row=start_row + 1, column=1)
