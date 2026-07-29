"""Generate client-downloadable CSV and Excel artifacts from structured results."""

from __future__ import annotations

import csv
import json
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from modules.models import Finding


HEADER_FILL = PatternFill("solid", fgColor="176B52")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def generate_report_exports(
    audit_id: str,
    findings: Iterable[Finding],
    summary: dict[str, Any],
    output_dir: Path,
) -> list[Path]:
    rows = list(findings)
    directory = Path(output_dir)
    directory.mkdir(parents=True, exist_ok=True)

    csv_path = directory / f"seo-audit-{audit_id}.csv"
    with csv_path.open("w", newline="", encoding="utf-8-sig") as stream:
        writer = csv.writer(stream)
        writer.writerow(
            [
                "Severity",
                "Category",
                "Issue Type",
                "Page URL",
                "Resource URL",
                "Evidence",
                "Recommendation",
                "Source File",
            ]
        )
        for finding in rows:
            writer.writerow(
                [
                    finding.severity.value,
                    finding.category,
                    finding.issue_type,
                    finding.page_url,
                    finding.resource_url,
                    finding.evidence,
                    finding.recommendation,
                    finding.source_file,
                ]
            )

    excel_path = directory / f"seo-audit-{audit_id}.xlsx"
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Executive Summary"
    summary_sheet.append(["Metric", "Value"])
    _style_header(summary_sheet)
    for key in (
        "finding_count",
        "target_url",
        "target_location",
        "page_limit",
        "artifact_count",
    ):
        if key in summary:
            summary_sheet.append([key.replace("_", " ").title(), summary[key]])
    for key in ("severity_counts", "category_counts", "semrush", "keyword_metrics"):
        if summary.get(key):
            summary_sheet.append(
                [key.replace("_", " ").title(), json.dumps(summary[key], default=str)]
            )
    summary_sheet.column_dimensions["A"].width = 28
    summary_sheet.column_dimensions["B"].width = 80

    findings_sheet = workbook.create_sheet("Technical Findings")
    findings_sheet.append(
        [
            "Severity",
            "Category",
            "Issue Type",
            "Page URL",
            "Resource URL",
            "Evidence",
            "Recommendation",
            "Source File",
        ]
    )
    _style_header(findings_sheet)
    for finding in rows:
        findings_sheet.append(
            [
                finding.severity.value,
                finding.category,
                finding.issue_type,
                finding.page_url,
                finding.resource_url,
                finding.evidence,
                finding.recommendation,
                finding.source_file,
            ]
        )
        row = findings_sheet.max_row
        for column in (4, 5):
            cell = findings_sheet.cell(row=row, column=column)
            if cell.value:
                cell.hyperlink = str(cell.value)
                cell.style = "Hyperlink"
    _format_tabular_sheet(findings_sheet)

    recommendations = summary.get("content_recommendations") or []
    if recommendations:
        content_sheet = workbook.create_sheet("Content Recommendations")
        headers = [
            "URL",
            "Current Title",
            "Proposed Title",
            "Current H1",
            "Proposed H1",
            "Current Meta Description",
            "Proposed Meta Description",
            "Human Review Required",
        ]
        content_sheet.append(headers)
        _style_header(content_sheet)
        for item in recommendations:
            content_sheet.append(
                [
                    item.get("url"),
                    item.get("title"),
                    item.get("proposed_title"),
                    item.get("h1"),
                    item.get("proposed_h1"),
                    item.get("meta_description"),
                    item.get("proposed_meta_description"),
                    item.get("requires_human_review", True),
                ]
            )
        _format_tabular_sheet(content_sheet)

    experiences = summary.get("page_experience") or []
    if experiences:
        experience_sheet = workbook.create_sheet("Performance & Accessibility")
        experience_sheet.append(
            [
                "URL",
                "Performance Score",
                "Accessibility Score",
                "Metrics",
                "Accessibility Issues",
            ]
        )
        _style_header(experience_sheet)
        for item in experiences:
            experience_sheet.append(
                [
                    item.get("url"),
                    item.get("performance_score"),
                    item.get("accessibility_score"),
                    json.dumps(item.get("metrics", {}), default=str),
                    json.dumps(item.get("accessibility_issues", []), default=str),
                ]
            )
        _format_tabular_sheet(experience_sheet)

    workbook.save(excel_path)
    return [csv_path, excel_path]


def _style_header(sheet) -> None:
    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions


def _format_tabular_sheet(sheet) -> None:
    for column in sheet.columns:
        letter = column[0].column_letter
        sheet.column_dimensions[letter].width = min(
            60,
            max(14, max(len(str(cell.value or "")) for cell in column) + 2),
        )
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
