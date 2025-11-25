import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment
import os
from datetime import datetime
from typing import Optional, TYPE_CHECKING

if TYPE_CHECKING:
    from modules.agent import SEOAgent

class ExcelReportBuilder:
    def __init__(self, template_path="SCD - Kahuina SEO Report.xlsx", output_path="SEO_Report_Generated.xlsx", agent: Optional["SEOAgent"] = None):
        self.template_path = template_path
        self.output_path = output_path
        self.wb = None
        self.agent = agent  # SEOAgent instance for AI-powered suggestions

    def load_workbook(self):
        if not os.path.exists(self.template_path):
            print(f"Error: Template file '{self.template_path}' not found.")
            # For development purposes, create a blank workbook if template is missing
            # But in production, we really need the template.
            self.wb = openpyxl.Workbook()
            print("Created a new blank workbook instead.")
        else:
            self.wb = openpyxl.load_workbook(self.template_path)

    def save_workbook(self):
        if self.wb:
            self.wb.save(self.output_path)
            print(f"Report saved to: {self.output_path}")

    def update_technical_seo_tab(self, crawl_output_dir: str):
        """
        Updates Tab A: "Technical SEO"
        """
        if not self.wb: return
        
        # Get the sheet - assuming name is "Technical SEO" or similar based on description
        # If name differs, code might need adjustment. Checking for "Technical SEO"
        sheet_name = "Technical SEO"
        if sheet_name not in self.wb.sheetnames:
             # Try to find it or create it
             if "Technical" in self.wb.sheetnames:
                 sheet_name = "Technical"
             else:
                 print(f"Warning: Sheet '{sheet_name}' not found. Creating it.")
                 self.wb.create_sheet(sheet_name)
        
        ws = self.wb[sheet_name]
        
        # Load CSVs
        files = {
            "4xx": os.path.join(crawl_output_dir, "response_codes_client_error_4xx.csv"),
            "missing_alt": os.path.join(crawl_output_dir, "images_missing_alt_text.csv"),
            # Add others if needed for other metrics
        }
        
        counts = {}
        for key, path in files.items():
            if os.path.exists(path):
                try:
                    df = pd.read_csv(path)
                    # Screaming Frog CSVs often have a header row 2 (index 1) or 1 (index 0). 
                    # Assuming standard export.
                    counts[key] = len(df)
                except Exception as e:
                    print(f"Error reading {path}: {e}")
                    counts[key] = 0
            else:
                counts[key] = 0
        
        # Logic: "Count rows -> Write count to 'Occurrences' column"
        # We need to find the "Occurrences" column and the row for "Broken Links" / "Missing Alt Text"
        # This requires knowing the exact cell positions or searching for labels.
        # Implementation strategy: Search for labels in column A/B.
        
        labels_map = {
            "Broken Links": counts.get("4xx", 0),
            "Missing Alt Text": counts.get("missing_alt", 0)
        }
        
        # Naive search in first 50 rows, 10 cols
        updated = False
        for row in ws.iter_rows(min_row=1, max_row=50, min_col=1, max_col=10):
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    val = cell.value.strip()
                    if val in labels_map:
                        # Assume "Occurrences" is in the next column or specific offset?
                        # Often templates have: Metric | Occurrences | ...
                        # Let's look for a number or empty cell in the next column
                        target_cell = ws.cell(row=cell.row, column=cell.column + 1) 
                        # Or verify header. For now, put it in next col
                        target_cell.value = labels_map[val]
                        updated = True
                        
                        # Update "Date Discovered" if present (maybe +2 cols?)
                        # Just an example update
                        
        if not updated:
            print("Warning: Could not find 'Broken Links' or 'Missing Alt Text' labels in Technical SEO tab.")


    def _get_valid_urls_from_crawl(self, crawl_output_dir: str) -> list:
        """
        Extract valid URLs from the internal_all.csv for use in 404 fix suggestions.
        """
        valid_urls = []
        internal_csv = os.path.join(crawl_output_dir, "internal_all.csv")
        if os.path.exists(internal_csv):
            try:
                df = pd.read_csv(internal_csv)
                if 'Address' in df.columns:
                    valid_urls = df['Address'].dropna().tolist()
            except Exception as e:
                print(f"Error reading internal URLs: {e}")
        return valid_urls

    def _load_page_data_from_crawl(self, crawl_output_dir: str) -> dict:
        """
        Load page metadata (Title, H1, Meta Description, etc.) from internal_all.csv
        Returns a dict keyed by URL with all available page data.
        
        Screaming Frog typical columns:
        - Address (URL)
        - Title 1, Title 1 Length, Title 1 Pixel Width
        - Meta Description 1, Meta Description 1 Length
        - H1-1, H1-2 (multiple H1s if present)
        - H2-1, H2-2
        - Canonical Link Element 1
        - Word Count
        - Indexability, Indexability Status
        """
        page_data = {}
        internal_csv = os.path.join(crawl_output_dir, "internal_all.csv")
        
        if os.path.exists(internal_csv):
            try:
                df = pd.read_csv(internal_csv)
                
                # Build lookup dict by URL
                for _, row in df.iterrows():
                    url = row.get('Address', '')
                    if not url:
                        continue
                    
                    page_data[url] = {
                        'url': url,
                        'title': self._safe_get(row, ['Title 1', 'Title']),
                        'title_length': self._safe_get(row, ['Title 1 Length', 'Title Length']),
                        'meta_description': self._safe_get(row, ['Meta Description 1', 'Meta Description']),
                        'meta_description_length': self._safe_get(row, ['Meta Description 1 Length', 'Meta Description Length']),
                        'h1': self._safe_get(row, ['H1-1', 'H1']),
                        'h1_2': self._safe_get(row, ['H1-2']),  # Second H1 if multiple
                        'h2': self._safe_get(row, ['H2-1', 'H2']),
                        'canonical': self._safe_get(row, ['Canonical Link Element 1', 'Canonical']),
                        'word_count': self._safe_get(row, ['Word Count']),
                        'indexability': self._safe_get(row, ['Indexability']),
                        'status_code': self._safe_get(row, ['Status Code']),
                        'content_type': self._safe_get(row, ['Content Type']),
                    }
            except Exception as e:
                print(f"Error loading page data from crawl: {e}")
        
        return page_data

    def _safe_get(self, row, column_names: list) -> str:
        """
        Safely get a value from a row, trying multiple possible column names.
        Returns empty string if not found or NaN.
        """
        for col in column_names:
            if col in row.index:
                val = row.get(col, '')
                if pd.notna(val):
                    return str(val)
        return ''

    def create_detailed_audit_tab(self, crawl_output_dir: str):
        """
        Creates Tab B: "Detailed Audit Logs"
        If an agent is provided, uses AI to generate specific fix suggestions.
        Now includes current page data (Title, H1, Meta Description) for context.
        """
        if not self.wb: return
        
        sheet_name = "Detailed Audit Logs"
        if sheet_name in self.wb.sheetnames:
            ws = self.wb[sheet_name]
            self.wb.remove(ws)
            ws = self.wb.create_sheet(sheet_name)
        else:
            ws = self.wb.create_sheet(sheet_name)
            
        # Enhanced Headers - include current page structure data
        headers = [
            "Issue Type", 
            "Page URL", 
            "Current Title",
            "Current H1",
            "Current Meta Description",
            "Element/Details", 
            "Suggested Fix"
        ]
        ws.append(headers)
        
        # Style headers
        for cell in ws[1]:
            cell.font = Font(bold=True)
        
        # Set column widths for better readability
        ws.column_dimensions['A'].width = 20  # Issue Type
        ws.column_dimensions['B'].width = 50  # Page URL
        ws.column_dimensions['C'].width = 40  # Current Title
        ws.column_dimensions['D'].width = 40  # Current H1
        ws.column_dimensions['E'].width = 50  # Current Meta Description
        ws.column_dimensions['F'].width = 40  # Element/Details
        ws.column_dimensions['G'].width = 60  # Suggested Fix
        
        # Get valid URLs for 404 suggestions
        valid_urls = self._get_valid_urls_from_crawl(crawl_output_dir) if self.agent else []
        
        # Load page data from internal_all.csv for cross-referencing
        page_data = self._load_page_data_from_crawl(crawl_output_dir)
        print(f"  Loaded page data for {len(page_data)} URLs from crawl...")
        
        # Collect all issues first for batch processing
        all_issues = {
            '404': [],
            '3xx': [],
            'missing_alt': [],
            'missing_alt_attr': [],
            'missing_h1': [],
            'multiple_h1': [],
            'multiple_h2': [],
            'missing_title': [],
            'short_title': [],
            'missing_meta': [],
            'missing_canonical': [],
            'security': [],
            'url_params': [],
            'nofollow': []
        }
        
        # Helper function to enrich issue with page data
        def enrich_with_page_data(issue: dict, url: str) -> dict:
            """Add current title, H1, meta description from page_data lookup."""
            if url in page_data:
                pd_entry = page_data[url]
                issue['current_title'] = pd_entry.get('title', '')
                issue['current_h1'] = pd_entry.get('h1', '')
                issue['current_meta_desc'] = pd_entry.get('meta_description', '')
            else:
                # URL not in internal_all.csv - might be external or 404
                issue.setdefault('current_title', '')
                issue.setdefault('current_h1', '')
                issue.setdefault('current_meta_desc', '')
            return issue
            
        # ============================================================
        # COLLECT ALL ISSUES
        # ============================================================
        
        # Collect 404s
        path_4xx = os.path.join(crawl_output_dir, "response_codes_client_error_(4xx).csv")
        if os.path.exists(path_4xx):
            try:
                df = pd.read_csv(path_4xx)
                for _, row in df.iterrows():
                    destination = row.get('Address', '')
                    source = row.get('Source', 'See Inlinks Report')
                    issue = {
                        'issue_type': '404 Error',
                        'page_url': source,
                        'element': f"Linked to: {destination}",
                        'broken_url': destination,
                        'source_url': source,
                        'suggested_fix': 'Update link or remove'
                    }
                    all_issues['404'].append(enrich_with_page_data(issue, source))
            except Exception as e:
                print(f"Error processing 4xx csv: {e}")

        # Collect 3xx Redirects
        path_3xx = os.path.join(crawl_output_dir, "response_codes_redirection_(3xx).csv")
        if os.path.exists(path_3xx):
            try:
                df = pd.read_csv(path_3xx)
                for _, row in df.iterrows():
                    source = row.get('Address', '')
                    dest = row.get('Redirect URI', '')
                    # Also try to get title/h1 from the CSV if available
                    title = self._safe_get(row, ['Title 1', 'Title'])
                    h1 = self._safe_get(row, ['H1-1', 'H1'])
                    meta_desc = self._safe_get(row, ['Meta Description 1', 'Meta Description'])
                    
                    issue = {
                        'issue_type': 'Redirect (3xx)',
                        'page_url': source,
                        'element': f"Redirects to: {dest}",
                        'source_url': source,
                        'redirect_destination': dest,
                        'current_title': title,
                        'current_h1': h1,
                        'current_meta_desc': meta_desc,
                        'suggested_fix': 'Check if permanent/necessary'
                    }
                    # Enrich if CSV didn't have the data
                    if not title and not h1:
                        issue = enrich_with_page_data(issue, source)
                    all_issues['3xx'].append(issue)
            except Exception as e:
                print(f"Error processing 3xx csv: {e}")

        # Collect Missing Alt Text
        path_alt = os.path.join(crawl_output_dir, "images_missing_alt_text.csv")
        if os.path.exists(path_alt):
            try:
                df = pd.read_csv(path_alt)
                for _, row in df.iterrows():
                    source = row.get('Source', row.get('Address', ''))
                    image = row.get('Address', '')
                    issue = {
                        'issue_type': 'Missing Alt',
                        'page_url': source,
                        'element': f"Image: {image}",
                        'image_url': image,
                        'suggested_fix': 'Add descriptive alt text'
                    }
                    all_issues['missing_alt'].append(enrich_with_page_data(issue, source))
            except Exception as e:
                print(f"Error processing missing alt csv: {e}")
        
        # Collect Missing Alt Attribute
        path_alt_attr = os.path.join(crawl_output_dir, "images_missing_alt_attribute.csv")
        if os.path.exists(path_alt_attr):
            try:
                df = pd.read_csv(path_alt_attr)
                for _, row in df.iterrows():
                    source = row.get('Source', row.get('Address', ''))
                    image = row.get('Address', '')
                    issue = {
                        'issue_type': 'Missing Alt Attribute',
                        'page_url': source,
                        'element': f"Image: {image}",
                        'image_url': image,
                        'suggested_fix': 'Add alt attribute'
                    }
                    all_issues['missing_alt_attr'].append(enrich_with_page_data(issue, source))
            except Exception as e:
                print(f"Error processing missing alt attr csv: {e}")

        # Collect Missing H1
        path_h1 = os.path.join(crawl_output_dir, "h1_missing.csv")
        if os.path.exists(path_h1):
            try:
                df = pd.read_csv(path_h1)
                for _, row in df.iterrows():
                    address = row.get('Address', '')
                    # Get title/meta from the CSV row if available
                    title = self._safe_get(row, ['Title 1', 'Title'])
                    meta_desc = self._safe_get(row, ['Meta Description 1', 'Meta Description'])
                    
                    issue = {
                        'issue_type': 'Missing H1',
                        'page_url': address,
                        'element': '(No H1 found on page)',
                        'current_title': title,
                        'current_h1': '',  # It's missing, that's the issue
                        'current_meta_desc': meta_desc,
                        'suggested_fix': 'Add H1 heading'
                    }
                    # Enrich if CSV didn't have title/meta
                    if not title:
                        issue = enrich_with_page_data(issue, address)
                        issue['current_h1'] = ''  # Ensure H1 stays empty
                    all_issues['missing_h1'].append(issue)
            except Exception as e:
                print(f"Error processing missing h1 csv: {e}")

        # Collect Multiple H1
        path_h1_multi = os.path.join(crawl_output_dir, "h1_multiple.csv")
        if os.path.exists(path_h1_multi):
            try:
                df = pd.read_csv(path_h1_multi)
                for _, row in df.iterrows():
                    address = row.get('Address', '')
                    # Get all H1s from the CSV
                    h1_1 = self._safe_get(row, ['H1-1', 'H1'])
                    h1_2 = self._safe_get(row, ['H1-2'])
                    h1_count = self._safe_get(row, ['H1 Count', 'H1-1 Count'])
                    title = self._safe_get(row, ['Title 1', 'Title'])
                    meta_desc = self._safe_get(row, ['Meta Description 1', 'Meta Description'])
                    
                    h1_display = h1_1
                    if h1_2:
                        h1_display = f"{h1_1} | {h1_2}"
                    
                    issue = {
                        'issue_type': 'Multiple H1',
                        'page_url': address,
                        'element': f"H1 Count: {h1_count}" if h1_count else f"Multiple H1s found",
                        'current_title': title,
                        'current_h1': h1_display,
                        'current_meta_desc': meta_desc,
                        'suggested_fix': 'Use only one H1 per page'
                    }
                    if not title:
                        issue = enrich_with_page_data(issue, address)
                        issue['current_h1'] = h1_display  # Keep the multiple H1s display
                    all_issues['multiple_h1'].append(issue)
            except Exception as e:
                print(f"Error processing multiple h1 csv: {e}")
                
        # Collect Multiple H2
        path_h2_multi = os.path.join(crawl_output_dir, "h2_multiple.csv")
        if os.path.exists(path_h2_multi):
            try:
                df = pd.read_csv(path_h2_multi)
                for _, row in df.iterrows():
                    address = row.get('Address', '')
                    h2_count = self._safe_get(row, ['H2 Count', 'H2-1 Count'])
                    issue = {
                        'issue_type': 'Multiple H2',
                        'page_url': address,
                        'element': f"H2 Count: {h2_count}" if h2_count else "Multiple H2s found",
                        'suggested_fix': 'Review H2 hierarchy'
                    }
                    all_issues['multiple_h2'].append(enrich_with_page_data(issue, address))
            except Exception as e:
                print(f"Error processing multiple h2 csv: {e}")

        # Collect Missing Page Titles
        path_titles = os.path.join(crawl_output_dir, "page_titles_missing.csv")
        if os.path.exists(path_titles):
            try:
                df = pd.read_csv(path_titles)
                for _, row in df.iterrows():
                    address = row.get('Address', '')
                    h1 = self._safe_get(row, ['H1-1', 'H1'])
                    meta_desc = self._safe_get(row, ['Meta Description 1', 'Meta Description'])
                    
                    issue = {
                        'issue_type': 'Missing Page Title',
                        'page_url': address,
                        'element': '(No title tag found)',
                        'current_title': '',  # It's missing
                        'current_h1': h1,
                        'current_meta_desc': meta_desc,
                        'suggested_fix': 'Add unique page title'
                    }
                    if not h1:
                        issue = enrich_with_page_data(issue, address)
                        issue['current_title'] = ''  # Keep title empty
                    all_issues['missing_title'].append(issue)
            except Exception as e:
                print(f"Error processing missing titles csv: {e}")
                
        # Collect Short Page Titles
        path_short_titles = os.path.join(crawl_output_dir, "page_titles_below_30_characters.csv")
        if os.path.exists(path_short_titles):
            try:
                df = pd.read_csv(path_short_titles)
                for _, row in df.iterrows():
                    address = row.get('Address', '')
                    title = self._safe_get(row, ['Title 1', 'Title'])
                    title_len = self._safe_get(row, ['Title 1 Length', 'Title Length'])
                    h1 = self._safe_get(row, ['H1-1', 'H1'])
                    meta_desc = self._safe_get(row, ['Meta Description 1', 'Meta Description'])
                    
                    issue = {
                        'issue_type': 'Short Page Title',
                        'page_url': address,
                        'element': f"Length: {title_len} chars" if title_len else f"Title too short",
                        'current_title': title,
                        'current_h1': h1,
                        'current_meta_desc': meta_desc,
                        'suggested_fix': 'Expand title (30-60 chars)'
                    }
                    all_issues['short_title'].append(issue)
            except Exception as e:
                print(f"Error processing short titles csv: {e}")

        # Collect Missing Meta Descriptions
        path_meta = os.path.join(crawl_output_dir, "meta_description_missing.csv")
        if os.path.exists(path_meta):
            try:
                df = pd.read_csv(path_meta)
                for _, row in df.iterrows():
                    address = row.get('Address', '')
                    title = self._safe_get(row, ['Title 1', 'Title'])
                    h1 = self._safe_get(row, ['H1-1', 'H1'])
                    
                    issue = {
                        'issue_type': 'Missing Meta Description',
                        'page_url': address,
                        'element': '(No meta description found)',
                        'current_title': title,
                        'current_h1': h1,
                        'current_meta_desc': '',  # It's missing
                        'suggested_fix': 'Add meta description (150-160 chars)'
                    }
                    if not title:
                        issue = enrich_with_page_data(issue, address)
                        issue['current_meta_desc'] = ''  # Keep meta empty
                    all_issues['missing_meta'].append(issue)
            except Exception as e:
                print(f"Error processing missing meta desc csv: {e}")
                
        # Collect Missing Canonicals
        path_canon = os.path.join(crawl_output_dir, "canonicals_missing.csv")
        if os.path.exists(path_canon):
            try:
                df = pd.read_csv(path_canon)
                for _, row in df.iterrows():
                    address = row.get('Address', '')
                    issue = {
                        'issue_type': 'Missing Canonical',
                        'page_url': address,
                        'element': '(No canonical tag found)',
                        'suggested_fix': 'Add self-referencing canonical'
                    }
                    all_issues['missing_canonical'].append(enrich_with_page_data(issue, address))
            except Exception as e:
                print(f"Error processing missing canonicals csv: {e}")

        # Collect Security Header Issues
        security_files = {
            "Missing HSTS": "security_missing_hsts.csv",
            "Missing X-Frame-Options": "security_missing_x-frame-options_header.csv",
            "Missing X-Content-Type-Options": "security_missing_x-content-type-options_header.csv",
            "Missing Referrer-Policy": "security_missing_secure_referrer-policy_header.csv",
            "Missing CSP": "security_missing_content-security-policy_header.csv"
        }
        for issue_name, filename in security_files.items():
            path = os.path.join(crawl_output_dir, filename)
            if os.path.exists(path):
                try:
                    df = pd.read_csv(path)
                    for _, row in df.iterrows():
                        address = row.get('Address', '')
                        issue = {
                            'issue_type': issue_name,
                            'page_url': address,
                            'element': '',
                            'suggested_fix': 'Add security header'
                        }
                        all_issues['security'].append(enrich_with_page_data(issue, address))
                except Exception as e:
                    print(f"Error processing {filename}: {e}")

        # Collect URL Parameters
        path_params = os.path.join(crawl_output_dir, "url_parameters.csv")
        if os.path.exists(path_params):
            try:
                df = pd.read_csv(path_params)
                for _, row in df.iterrows():
                    address = row.get('Address', '')
                    issue = {
                        'issue_type': 'URL Parameters',
                        'page_url': address,
                        'element': '',
                        'suggested_fix': 'Check for duplicate content'
                    }
                    all_issues['url_params'].append(enrich_with_page_data(issue, address))
            except Exception as e:
                print(f"Error processing url params csv: {e}")

        # Collect External Nofollow Links
        path_ext = os.path.join(crawl_output_dir, "links_external.csv")
        if os.path.exists(path_ext):
            try:
                df = pd.read_csv(path_ext)
                if 'Follow' in df.columns:
                    nofollows = df[df['Follow'] == False]
                    if nofollows.empty and df['Follow'].dtype == object:
                        nofollows = df[df['Follow'].astype(str).str.lower() == 'false']
                    
                    for _, row in nofollows.iterrows():
                        source = row.get('Source', '')
                        dest = row.get('Destination', row.get('Address', ''))
                        issue = {
                            'issue_type': 'External Nofollow',
                            'page_url': source,
                            'element': f"Link to: {dest}",
                            'suggested_fix': 'Verify nofollow is intended'
                        }
                        all_issues['nofollow'].append(enrich_with_page_data(issue, source))
            except Exception as e:
                print(f"Error processing external links csv: {e}")

        # ============================================================
        # PROCESS WITH AI AGENT (if available)
        # ============================================================
        
        if self.agent:
            print("\n  Using AI agent to generate specific fix suggestions...")
            
            # Process 404s with AI
            if all_issues['404']:
                print(f"  Processing {len(all_issues['404'])} 404 errors...")
                all_issues['404'] = self.agent.suggest_404_fixes_batch(all_issues['404'], valid_urls)
            
            # Process redirects with AI
            if all_issues['3xx']:
                print(f"  Processing {len(all_issues['3xx'])} redirects...")
                all_issues['3xx'] = self.agent.suggest_redirect_fixes_batch(all_issues['3xx'])
            
            # Process missing alt text with AI
            if all_issues['missing_alt']:
                print(f"  Processing {len(all_issues['missing_alt'])} missing alt text issues...")
                all_issues['missing_alt'] = self.agent.generate_alt_text_batch(all_issues['missing_alt'])
            
            # Process missing alt attribute with AI (same as missing alt)
            if all_issues['missing_alt_attr']:
                print(f"  Processing {len(all_issues['missing_alt_attr'])} missing alt attribute issues...")
                all_issues['missing_alt_attr'] = self.agent.generate_alt_text_batch(all_issues['missing_alt_attr'])
            
            # Process missing H1 with AI
            if all_issues['missing_h1']:
                print(f"  Processing {len(all_issues['missing_h1'])} missing H1 issues...")
                all_issues['missing_h1'] = self.agent.generate_h1_suggestions_batch(all_issues['missing_h1'])
            
            # Process missing titles with AI
            if all_issues['missing_title']:
                print(f"  Processing {len(all_issues['missing_title'])} missing title issues...")
                all_issues['missing_title'] = self.agent.generate_title_suggestions_batch(all_issues['missing_title'])
            
            # Process short titles with AI
            if all_issues['short_title']:
                print(f"  Processing {len(all_issues['short_title'])} short title issues...")
                all_issues['short_title'] = self.agent.generate_title_suggestions_batch(all_issues['short_title'])
            
            # Process missing meta descriptions with AI
            if all_issues['missing_meta']:
                print(f"  Processing {len(all_issues['missing_meta'])} missing meta description issues...")
                all_issues['missing_meta'] = self.agent.generate_meta_description_batch(all_issues['missing_meta'])
            
            # Process missing canonicals with AI
            if all_issues['missing_canonical']:
                print(f"  Processing {len(all_issues['missing_canonical'])} missing canonical issues...")
                all_issues['missing_canonical'] = self.agent.suggest_canonical_fixes_batch(all_issues['missing_canonical'])
        
        # ============================================================
        # WRITE ALL ISSUES TO WORKSHEET
        # ============================================================
        
        for issue_type, issues in all_issues.items():
            for issue in issues:
                ws.append([
                    issue.get('issue_type', ''),
                    issue.get('page_url', ''),
                    issue.get('current_title', ''),
                    issue.get('current_h1', ''),
                    issue.get('current_meta_desc', ''),
                    issue.get('element', ''),
                    issue.get('suggested_fix', '')
                ])
        
        # Apply text wrapping to content columns
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical='top')


    def update_onpage_recommendations(self, recommendations: list):
        """
        Updates Tab C: "On-Page Recommendations"
        recommendations: list of dicts with url, keyword, and content data
        Enhanced to show current title, H1, and meta description separately.
        """
        if not self.wb: return
        sheet_name = "On-Page Recommendations"
        
        # Remove existing sheet if it exists (handles merged cells issue)
        if sheet_name in self.wb.sheetnames:
            self.wb.remove(self.wb[sheet_name])
        
        # Create fresh sheet
        ws = self.wb.create_sheet(sheet_name)
        
        # Set column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 80
        
        current_row = 1
            
        for rec in recommendations:
            # Web Page URL
            ws.cell(row=current_row, column=1, value="Web Page:")
            ws.cell(row=current_row, column=2, value=rec.get('url', ''))
            
            # Targeted Keyword
            ws.cell(row=current_row+1, column=1, value="Targeted Keyword:")
            ws.cell(row=current_row+1, column=2, value=rec.get('keyword', ''))
            
            # Current Title
            ws.cell(row=current_row+2, column=1, value="Current Title:")
            ws.cell(row=current_row+2, column=2, value=rec.get('current_title', ''))
            
            # Proposed Title
            ws.cell(row=current_row+3, column=1, value="Proposed Title:")
            ws.cell(row=current_row+3, column=2, value=rec.get('proposed_title', ''))
            
            # Current H1
            ws.cell(row=current_row+4, column=1, value="Current H1:")
            ws.cell(row=current_row+4, column=2, value=rec.get('current_h1', ''))
            
            # Proposed H1
            ws.cell(row=current_row+5, column=1, value="Proposed H1:")
            ws.cell(row=current_row+5, column=2, value=rec.get('proposed_h1', ''))
            
            # Current Meta Description
            ws.cell(row=current_row+6, column=1, value="Current Meta Desc:")
            ws.cell(row=current_row+6, column=2, value=rec.get('current_meta_desc', ''))
            
            # Proposed Meta Description
            ws.cell(row=current_row+7, column=1, value="Proposed Meta Desc:")
            ws.cell(row=current_row+7, column=2, value=rec.get('proposed_meta_desc', ''))
            
            # Original Content (full block if provided)
            if rec.get('original_content'):
                ws.cell(row=current_row+8, column=1, value="Original Copy:")
                ws.cell(row=current_row+8, column=2, value=rec.get('original_content', ''))
                
                ws.cell(row=current_row+9, column=1, value="Proposed Copy:")
                ws.cell(row=current_row+9, column=2, value=rec.get('proposed_content', ''))
                rows_used = 11
            else:
                rows_used = 9
            
            # Simple styling
            for r in range(current_row, current_row + rows_used - 1):
                ws.cell(row=r, column=1).font = Font(bold=True)
                ws.cell(row=r, column=2).alignment = Alignment(wrap_text=True)
            
            current_row += rows_used

    def update_metadata_tab(self, metadata_list: list):
        """
        Updates Tab D: "Title Tags" & "Meta Descriptions" (or combined)
        metadata_list: list of dicts with url, keywords, current/proposed title and description
        Enhanced to show all current page data clearly.
        """
        if not self.wb: return
        
        # Create or use "Metadata Optimization" sheet for comprehensive view
        sheet_name = "Metadata Optimization"
        
        if sheet_name in self.wb.sheetnames:
            ws = self.wb[sheet_name]
            # Clear existing content
            self.wb.remove(ws)
            ws = self.wb.create_sheet(sheet_name)
        else:
            ws = self.wb.create_sheet(sheet_name)
        
        # Enhanced headers
        headers = [
            "URL", 
            "Target Keywords",
            "Current Title", 
            "Title Length",
            "Proposed Title", 
            "Proposed Title Length",
            "Current H1",
            "Current Meta Description", 
            "Meta Desc Length",
            "Proposed Meta Description", 
            "Proposed Meta Desc Length"
        ]
        ws.append(headers)
        
        # Style headers
        for cell in ws[1]:
            cell.font = Font(bold=True)
        
        # Set column widths
        ws.column_dimensions['A'].width = 50  # URL
        ws.column_dimensions['B'].width = 35  # Keywords
        ws.column_dimensions['C'].width = 45  # Current Title
        ws.column_dimensions['D'].width = 12  # Title Length
        ws.column_dimensions['E'].width = 45  # Proposed Title
        ws.column_dimensions['F'].width = 12  # Proposed Title Length
        ws.column_dimensions['G'].width = 40  # Current H1
        ws.column_dimensions['H'].width = 55  # Current Meta Desc
        ws.column_dimensions['I'].width = 12  # Meta Desc Length
        ws.column_dimensions['J'].width = 55  # Proposed Meta Desc
        ws.column_dimensions['K'].width = 12  # Proposed Meta Desc Length
        
        for item in metadata_list:
            current_title = item.get('current_title', '') or ''
            proposed_title = item.get('proposed_title', '') or ''
            current_desc = item.get('current_desc', '') or ''
            proposed_desc = item.get('proposed_desc', '') or ''
            current_h1 = item.get('current_h1', '') or ''
            
            ws.append([
                item.get('url', ''),
                item.get('keywords', ''),
                current_title,
                len(current_title),
                proposed_title,
                len(proposed_title),
                current_h1,
                current_desc,
                len(current_desc),
                proposed_desc,
                len(proposed_desc)
            ])
        
        # Apply text wrapping
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        # Also update "Title Tags" sheet if it exists (for backwards compatibility)
        if "Title Tags" in self.wb.sheetnames:
            ws_titles = self.wb["Title Tags"]
            start_row = ws_titles.max_row + 1
            for item in metadata_list:
                current_title = item.get('current_title', '') or ''
                proposed_title = item.get('proposed_title', '') or ''
                ws_titles.append([
                    item.get('url'),
                    item.get('keywords'),
                    current_title,
                    proposed_title,
                    len(proposed_title)
                ])


